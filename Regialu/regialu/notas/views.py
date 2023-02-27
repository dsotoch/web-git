from django.shortcuts import render
from aula.models import aula
from notas.models import notas
from area.models import area
from horario.models import horario
from alumnos.models import alumnos
from horario.models import horario
from django.http import JsonResponse, FileResponse, HttpResponse
import json
from login.models import usuario
from datetime import datetime
from django.contrib.auth.decorators import login_required
# PDF
from io import BytesIO
import xhtml2pdf.pisa as pisa
# EXCEL
import openpyxl
from openpyxl.styles import Font, PatternFill


@login_required
def index_notas(request):
    return render(request, 'indexNotas.html')


@login_required
def get_horario(request, id):
    try:
        user = usuario.objects.get(dni=request.user)

        classroom = aula.objects.get(idAula=id, user=user)
        schedules = horario.objects.filter(user=user,
                                           aulas=classroom, estado='ACT').prefetch_related('areas')
        dicc = []

        for n in schedules:

            ob = {'idHorario': n.idHorario, 'area': n.areas.descripcion, 'idArea': n.areas.idArea, 'turno': n.turno, 'dia': n.dia,
                  'horaInicio': str(n.horainicio), 'horaFin': str(n.horafin)}
            dicc.append(ob)
        return JsonResponse(json.dumps(dicc), safe=False)
    except Exception as ex:
        print(ex)
        return JsonResponse({'response': 'error'})


@login_required
def get_notes(request, idaula, idalumno, areas):
    try:
        if request.method == 'GET':
            user = usuario.objects.get(dni=request.user)

            tipo = request.GET.get('tipo')
            unidad = request.GET.get('unidad')
            classroom = aula.objects.get(idAula=idaula, user=user)
            students = alumnos.objects.get(
                idAlumno=idalumno, aula=classroom, user=user)
            area2 = area.objects.get(idArea=areas, user=user)
            notes = notas.objects.filter(user=user,
                                         alumno=students, area=area2, tipo=tipo, unidad=unidad)
            list = []
            for n in notes:
                dicc = {'idNota': n.idNota,
                        'descripcion': n.descripcion, 'calificativo': n.valor}
                list.append(dicc)

            return JsonResponse(json.dumps(list), safe=False)

    except Exception as ex:
        print(ex)
        return JsonResponse({'response': 'error'})


@login_required
def get_students(request, id):
    try:
        if request.method == 'GET':
            user = usuario.objects.get(dni=request.user)
            tipo = request.GET.get('tipo')
            unidad = request.GET.get('unidad')

            classroom = aula.objects.get(idAula=id, user=user)
            students = alumnos.objects.filter(aula=classroom, user=user)
            horarios = horario.objects.filter(user=user,
                                              aulas=classroom).prefetch_related('areas')[0]

            dicc = []
            for n in students:
                notes = notas.objects.filter(user=user,
                                             alumno=n, tipo=tipo, unidad=unidad, area=horarios.areas).count()

                stud = {'idAlumno': n.idAlumno,
                        'apellidos': n.apellidos, 'nombres': n.nombres, 'numeroNotas': notes}
                dicc.append(stud)
            return JsonResponse(json.dumps(dicc), safe=False)
    except Exception as ex:
        print(ex)
        return JsonResponse({'response', 'error'})


@login_required
def save_note(request):
    try:
        if request.method == 'POST':
            user = usuario.objects.get(dni=request.user)

            data = json.loads(request.body)
            description = data.get('description')
            note = data.get('note')
            classroom = data.get('classroom')
            student = data.get('student')
            id_area = data.get('area')
            tipo = data.get('tipo')
            unidad = data.get('unidad')
            areas = area.objects.get(idArea=id_area, user=user)
            object_classroom = aula.objects.get(idAula=classroom, user=user)
            object_student = alumnos.objects.get(idAlumno=student, user=user)
            notas.objects.create(user=user,
                                 tipo=tipo, unidad=unidad, descripcion=description, valor=note, alumno=object_student, area=areas)
            return JsonResponse({'response': 'success'})

    except Exception as ex:
        print(ex)
        return JsonResponse({'response': 'error'})


@login_required
def update_note(request, id):
    try:
        if request.method == 'PUT':
            user = usuario.objects.get(dni=request.user)

            data = json.loads(request.body)
            new_valor = data.get('new_value')
            object_note = notas.objects.get(idNota=id, user=user)
            object_note.valor = new_valor
            object_note.save()
            return JsonResponse({'response': 'success'})

    except Exception as ex:
        print(ex)
        return JsonResponse({'response': 'error'})


@login_required
def pdf_selected(request):

    try:
        if request.method == 'POST':
            user = usuario.objects.get(dni=request.user)

            data = json.loads(request.body)
            unidad = data.get('unidad')
            areas = data.get('areas')
            tipos = data.get('tipos')
            institucions = data.get('institucions')
            alumnoss = data.get('alumnos')
            periodo = data.get('periodo')
            fechas = datetime.now()
            object_area = area.objects.get(idArea=areas, user=user)
            object_alumno = alumnos.objects.get(idAlumno=alumnoss, user=user)
            notes = notas.objects.filter(user=user,
                                         alumno=object_alumno, area=object_area, tipo=tipos, unidad=unidad)
            numeroNotas = notas.objects.filter(user=user,
                                               alumno=object_alumno, area=object_area, tipo=tipos, unidad=unidad).count()
            promedio = 0
            promedios = 0
            for n in notes:
                promedios = int(n.valor)+promedios
                promedio = promedios/numeroNotas

            context = {'areas': object_area.descripcion, 'tipos': tipos, 'institucions': institucions, 'alumnoss': object_alumno.apellidos + " "+object_alumno.nombres,
                       'numeroNotas': numeroNotas, 'fechas': fechas, 'notes': notes, 'notas': numeroNotas, 'unidad': unidad, 'promedio': promedio, 'periodo': periodo}

            return render(request, 'pdfSelected.html', context)

    except Exception as ex:
        print(ex)
        return JsonResponse({'response': 'error'})


@login_required
def generate_pdf(request):
    html = request.POST.get('html', '')
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        response = HttpResponse(
            result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="example.pdf"'
        return response
    else:

        return HttpResponse("Error al generar el PDF")


@login_required
def export_to_excel(request):
    try:
        user = usuario.objects.get(dni=request.user)

        data = json.loads(request.body)
        name_aula = data.get('nombre_aula')
        unidad = data.get('unidad')
        areas = data.get('areas')
        idarea = data.get('idarea')
        tipos = data.get('tipos')
        institucions = data.get('institucions')
        alumnoss = data.get('alumnos')
        periodo = data.get('periodo')
        object_area = area.objects.get(idArea=idarea, user=user)
        # Crea una hoja de cálculo
        new_libro = openpyxl.Workbook()
        hoja = new_libro.active

        # Agrega algunos datos a la hoja de cálculo
        hoja['A1'] = 'Periodo'
        hoja['B1'] = 'Area'
        hoja['C1'] = 'Tipo'
        hoja['D1'] = 'Unidad'
        hoja['E1'] = 'Descripcion-Nota'
        hoja['F1'] = 'Calificación'
        hoja['G1'] = 'Apellidos'
        hoja['H1'] = 'Nombres'
        # Personalizar celdas
        font = Font(size=15, bold=True)

        fill = PatternFill(start_color='FFFF00',
                           end_color='FFFF00', fill_type='solid')

        hoja['A1'].font = font
        hoja['B1'].font = font
        hoja['C1'].font = font
        hoja['D1'].font = font
        hoja['E1'].font = font
        hoja['F1'].font = font
        hoja['G1'].font = font
        hoja['H1'].font = font
        hoja['A1'].fill = fill
        hoja['B1'].fill = fill
        hoja['C1'].fill = fill
        hoja['D1'].fill = fill
        hoja['E1'].fill = fill
        hoja['F1'].fill = fill
        hoja['G1'].fill = fill
        hoja['H1'].fill = fill
        # Define el ancho de la columna
        hoja.column_dimensions['A'].width = 40
        hoja.column_dimensions['B'].width = 30
        hoja.column_dimensions['C'].width = 30
        hoja.column_dimensions['E'].width = 40
        hoja.column_dimensions['F'].width = 25
        hoja.column_dimensions['G'].width = 40
        hoja.column_dimensions['H'].width = 40

        con = "2"
        for n in alumnoss:
            object_alumno = alumnos.objects.filter(idAlumno=n, user=user)

            for x in object_alumno:
                object_notas = notas.objects.filter(user=user,
                                                    alumno=x, area=object_area, unidad=unidad, tipo=tipos).prefetch_related('alumno')

                for i in object_notas:
                    alu = i.alumno
                    hoja["A"+con+""] = periodo
                    hoja["B"+con+""] = areas
                    hoja["C"+con+""] = tipos
                    hoja["D"+con+""] = unidad
                    hoja["E"+con+""] = i.descripcion
                    hoja["F"+con+""] = i.valor
                    hoja["G"+con+""] = alu.apellidos
                    hoja["H"+con+""] = alu.nombres

                    con = str(int(con)+1)

        # Guarda el archivo de Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="' +name_aula+'".xlsx'
        # Guarda el archivo en la respuesta HTTP
        new_libro.save(response)
        return response
    except Exception as es:
        print(es)
